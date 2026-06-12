# ============================================================================
# main.py — Punto de entrada principal del Auditor de Excel
# ============================================================================
"""
Script principal que:
  1. Valida la existencia de la PLANTILLA.xlsx y la carpeta TRABAJOS_ESTUDIANTES.
  2. Itera por todos los archivos .xlsx de la carpeta.
  3. Ejecuta la auditoría completa para cada archivo.
  4. Guarda los archivos revisados en la subcarpeta /REVISADOS.
  5. Genera el archivo LOG_NOTAS.csv con el resumen final.

Uso:
    python main.py
    python main.py --plantilla "ruta/a/PLANTILLA.xlsx" --trabajos "ruta/a/TRABAJOS_ESTUDIANTES"
"""

import os
import sys
import csv
import argparse
import logging
from datetime import datetime

from config import (
    PLANTILLA_PATH,
    TRABAJOS_DIR,
    REVISADOS_DIR,
    LOG_NOTAS_PATH,
    EXTENSIONES_VALIDAS,
)
from auditor import auditar_libro


# ============================================================================
# CONFIGURACIÓN DE LOGGING
# ============================================================================
def configurar_logging():
    """Configura el sistema de logging con salida básica a consola."""
    log_format = "%(asctime)s | %(levelname)-7s | %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"

    # Logger raíz
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Limpiar handlers existentes para evitar duplicación
    logger.handlers = []

    # Handler de consola
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter(log_format, datefmt=date_format))
    logger.addHandler(ch)

    return logger


def agregar_log_archivo(logger, seccion, fecha):
    """Crea la carpeta LOGS si no existe y añade el manejador de archivo de log."""
    log_format = "%(asctime)s | %(levelname)-7s | %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"
    
    logs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "LOGS")
    os.makedirs(logs_dir, exist_ok=True)
    
    log_filename = f"{seccion}_auditoria_{fecha}.log"
    log_file = os.path.join(logs_dir, log_filename)
    
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(log_format, datefmt=date_format))
    logger.addHandler(fh)
    
    return log_file


# ============================================================================
# VALIDACIONES DE ENTRADA
# ============================================================================
def validar_rutas(ruta_plantilla, ruta_trabajos):
    """
    Valida que la plantilla exista y la carpeta de trabajos exista y
    contenga archivos .xlsx.

    Retorna:
        list: Lista de rutas absolutas a los archivos .xlsx encontrados.

    Lanza:
        FileNotFoundError: Si la plantilla o la carpeta no existen.
        ValueError: Si no se encuentran archivos .xlsx.
    """
    # Validar plantilla
    if not os.path.isfile(ruta_plantilla):
        raise FileNotFoundError(
            f"❌ No se encontró la plantilla: '{ruta_plantilla}'\n"
            f"   Asegúrate de que el archivo PLANTILLA.xlsx exista en la ruta indicada."
        )

    # Validar carpeta de trabajos
    if not os.path.isdir(ruta_trabajos):
        raise FileNotFoundError(
            f"❌ No se encontró la carpeta de trabajos: '{ruta_trabajos}'\n"
            f"   Asegúrate de que la carpeta TRABAJOS_ESTUDIANTES exista."
        )

    # Buscar archivos .xlsx
    archivos = []
    for archivo in sorted(os.listdir(ruta_trabajos)):
        if archivo.lower().endswith(EXTENSIONES_VALIDAS) and not archivo.startswith("~$"):
            archivos.append(os.path.join(ruta_trabajos, archivo))

    if not archivos:
        raise ValueError(
            f"❌ No se encontraron archivos .xlsx en: '{ruta_trabajos}'\n"
            f"   Coloca los trabajos de los estudiantes en esa carpeta."
        )

    return archivos


# ============================================================================
# GENERACIÓN DEL LOG XLSX
# ============================================================================
def generar_log_xlsx(resultados, ruta_xlsx, ruta_plantilla, ruta_trabajos):
    """
    Genera el archivo LOG_NOTAS_fecha-hora.xlsx con el resumen detallado de la auditoría.
    
    Aplica formatos profesionales modernos:
      - Encabezado con color de relleno Azul Primario (#0056b3) y texto blanco.
      - Fila 1 de cabecera alta con wrap_text.
      - Columnas de porcentajes con fondo azul suave (#FFE6F0FA).
      - Rejilla fina en todos los datos.
      - Formato corto de fecha (dd/mm/yyyy) y valores numéricos nativos.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Obtener los nombres de las hojas de la plantilla para crear las columnas
    try:
        import unicodedata
        
        def es_rubrica(nombre):
            norm = "".join(
                c for c in unicodedata.normalize('NFD', nombre)
                if unicodedata.category(c) != 'Mn'
            ).lower().strip()
            return norm in ("rubrica", "rubricas")

        wb_p = openpyxl.load_workbook(ruta_plantilla, read_only=True)
        hojas_plantilla = [
            name for name in wb_p.sheetnames
            if wb_p[name].sheet_state == "visible" and not es_rubrica(name)
        ]
        wb_p.close()
    except Exception as e:
        logging.error(f"❌ Error al leer las hojas de la plantilla: {e}")
        hojas_plantilla = []

    # Extraer ID_Seccion del nombre del directorio de trabajos
    id_seccion = os.path.basename(os.path.abspath(ruta_trabajos))
    if not id_seccion:
        id_seccion = "SEC_DEFAULT"

    # Crear nuevo libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Auditoría"

    # Construir encabezados dinámicamente
    encabezados = [
        "Fecha",
        "ID_Estudiante"
    ]
    
    for hoja in hojas_plantilla:
        encabezados.extend([
            f"Aciertos {hoja}",
            f"Errores {hoja}",
            f"Porcentaje {hoja} (%)"
        ])
        
    encabezados.extend([
        "Aciertos Total",
        "Errores Total",
        "Total Evaluaciones",
        "Porcentaje Total (%)",
        "Errores en Objetos/COM"
    ])

    # Estilos del encabezado
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid") # Azul Primario #0056b3
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Configurar altura de la cabecera
    ws.row_dimensions[1].height = 42

    # Escribir cabecera
    for col_idx, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_idx, value=encabezado)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = header_align
        celda.border = thin_border

    # Escribir filas de datos
    fecha_actual_str = datetime.now().strftime("%d/%m/%Y")
    try:
        fecha_objeto = datetime.strptime(fecha_actual_str, "%d/%m/%Y").date()
    except Exception:
        fecha_objeto = fecha_actual_str

    # Estilos de datos
    data_font = Font(name="Calibri", size=11, color="000000")
    pct_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid") # Verde suave #E2EFDA
    pct_font = Font(name="Calibri", size=11, bold=True, color="375623") # Verde oscuro #375623
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_idx, res in enumerate(resultados, 2):
        id_estudiante = os.path.splitext(res["archivo"])[0]

        # Fila base
        ws.cell(row=row_idx, column=1, value=fecha_objeto).number_format = 'dd/mm/yyyy'
        ws.cell(row=row_idx, column=1).alignment = align_center
        
        ws.cell(row=row_idx, column=2, value=id_estudiante).alignment = align_left

        curr_col = 3

        # Detalle por hoja
        for hoja in hojas_plantilla:
            detalle_hoja = None
            for h in res.get("detalle_hojas", []):
                if h["hoja"] == hoja:
                    detalle_hoja = h
                    break

            if detalle_hoja is not None:
                ac_hoja = detalle_hoja["aciertos"]
                er_hoja = detalle_hoja["errores"]
            else:
                ac_hoja = 0
                er_hoja = 1 if res["total_errores"] != -1 else 0

            tot_hoja = ac_hoja + er_hoja
            pct_hoja = (ac_hoja / tot_hoja * 100) if tot_hoja > 0 else 0.0

            # Escribir aciertos
            c_ac = ws.cell(row=row_idx, column=curr_col, value=ac_hoja)
            c_ac.alignment = align_right
            c_ac.number_format = '#,##0'
            curr_col += 1

            # Escribir errores
            c_er = ws.cell(row=row_idx, column=curr_col, value=er_hoja)
            c_er.alignment = align_right
            c_er.number_format = '#,##0'
            curr_col += 1

            # Escribir porcentaje
            c_pct = ws.cell(row=row_idx, column=curr_col, value=float(pct_hoja))
            c_pct.alignment = align_right
            c_pct.number_format = '0.0'
            curr_col += 1

        # Totales del libro
        aciertos_tot = res["total_aciertos"]
        errores_tot = res["total_errores"]
        total_tot = aciertos_tot + errores_tot if errores_tot >= 0 else 0
        pct_tot = (aciertos_tot / total_tot * 100) if total_tot > 0 else 0.0

        errores_objetos = sum(
            len(h.get("detalles_objetos", []))
            for h in res.get("detalle_hojas", [])
        )
        errores_com = len(res.get("detalles_com", []))
        total_err_obj_com = errores_objetos + errores_com

        # Aciertos total
        c_act = ws.cell(row=row_idx, column=curr_col, value=aciertos_tot)
        c_act.alignment = align_right
        c_act.number_format = '#,##0'
        curr_col += 1

        # Errores total
        c_ert = ws.cell(row=row_idx, column=curr_col, value=errores_tot if errores_tot >= 0 else 0)
        c_ert.alignment = align_right
        c_ert.number_format = '#,##0'
        curr_col += 1

        # Total evaluaciones
        c_tte = ws.cell(row=row_idx, column=curr_col, value=total_tot)
        c_tte.alignment = align_right
        c_tte.number_format = '#,##0'
        curr_col += 1

        # Porcentaje total
        c_pctt = ws.cell(row=row_idx, column=curr_col, value=float(pct_tot))
        c_pctt.alignment = align_right
        c_pctt.number_format = '0.0'
        curr_col += 1

        # Errores objetos/com
        c_eoc = ws.cell(row=row_idx, column=curr_col, value=total_err_obj_com)
        c_eoc.alignment = align_right
        c_eoc.number_format = '#,##0'

        # Estilo de datos de toda la fila (borde, font y relleno condicional)
        for col_idx in range(1, len(encabezados) + 1):
            celda = ws.cell(row=row_idx, column=col_idx)
            celda.border = thin_border
            
            header_text = encabezados[col_idx - 1]
            if "porcentaje" in header_text.lower() or "%" in header_text:
                celda.font = pct_font
                celda.fill = pct_fill
            else:
                celda.font = data_font

    # Ajustar anchos de columna basados en los datos (fila 2+)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col[1:]: # Omitir fila 1
            if cell.value is not None:
                val_str = cell.value.strftime('%d/%m/%Y') if hasattr(cell.value, 'strftime') else str(cell.value)
                max_len = max(max_len, len(val_str))
        # Ancho mínimo de 15 para cabeceras con wrap_text
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Activar autofilter en el encabezado
    last_col = get_column_letter(len(encabezados))
    last_row = len(resultados) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    try:
        wb.save(ruta_xlsx)
        logging.info(f"\n📊 Reporte Excel generado en: {ruta_xlsx}")
    except Exception as e:
        logging.error(f"❌ No se pudo guardar el archivo Excel '{ruta_xlsx}': {e}")


# ============================================================================
# FLUJO PRINCIPAL
# ============================================================================
def main():
    """Punto de entrada principal del script de auditoría."""
    # Evitar crasheos de encoding por emojis unicode en consolas Windows
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass

    logger = configurar_logging()

    # Parsear argumentos opcionales
    parser = argparse.ArgumentParser(
        description="🔍 Auditor de Excel — Comparación de trabajos de estudiantes vs plantilla"
    )
    parser.add_argument(
        "--plantilla",
        default=PLANTILLA_PATH,
        help=f"Ruta al archivo PLANTILLA.xlsx (default: {PLANTILLA_PATH})"
    )
    parser.add_argument(
        "--trabajos",
        default=TRABAJOS_DIR,
        help=f"Ruta a la carpeta TRABAJOS_ESTUDIANTES (default: {TRABAJOS_DIR})"
    )
    parser.add_argument(
        "--salida",
        default=REVISADOS_DIR,
        help=f"Ruta a la carpeta de salida REVISADOS (default: {REVISADOS_DIR})"
    )
    parser.add_argument(
        "--log",
        default=None,
        help="Ruta al archivo de resultados de Excel (.xlsx)"
    )
    parser.add_argument(
        "--fecha",
        default=None,
        help="Fecha de creacion personalizada (ej: 11MAY26)"
    )
    parser.add_argument(
        "--seccion",
        default=None,
        help="Seccion del grupo evaluado (ej: 11-5B)"
    )
    args = parser.parse_args()

    # Determinar seccion y fecha para el formato de nombres de archivo
    seccion = args.seccion
    if not seccion:
        seccion = os.path.basename(os.path.abspath(args.trabajos))
        if not seccion or seccion == "TRABAJOS_ESTUDIANTES":
            seccion = "11-XB"

    fecha = args.fecha
    if not fecha:
        fecha = datetime.now().strftime("%d%m%y")

    # Crear carpeta de logs si no existe
    from config import LOGS_DIR
    os.makedirs(LOGS_DIR, exist_ok=True)

    # Agregar manejador de archivo de logs en la carpeta LOGS
    agregar_log_archivo(logger, seccion, fecha)

    # Configurar la ruta de salida del Excel final
    if args.log:
        args.log = os.path.abspath(args.log)
    else:
        args.log = os.path.join(LOGS_DIR, f"{seccion}_LOG_NOTAS_{fecha}.xlsx")

    # Banner
    logger.info("=" * 70)
    logger.info("  🔍 AUDITOR DE EXCEL — Comparación de Fidelidad Total")
    logger.info(f"  📅 Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 70)

    # Validar rutas
    try:
        archivos_estudiantes = validar_rutas(args.plantilla, args.trabajos)
    except (FileNotFoundError, ValueError) as e:
        logger.error(str(e))
        sys.exit(1)

    logger.info(f"\n📂 Plantilla:  {args.plantilla}")
    logger.info(f"📂 Trabajos:   {args.trabajos}")
    logger.info(f"📂 Salida:     {args.salida}")
    logger.info(f"📄 Archivos encontrados: {len(archivos_estudiantes)}")

    # Crear carpeta de salida
    os.makedirs(args.salida, exist_ok=True)

    # ----------------------------------------------------------------
    # ITERAR POR CADA ARCHIVO DE ESTUDIANTE
    # ----------------------------------------------------------------
    resultados = []
    total_archivos = len(archivos_estudiantes)

    for idx, ruta_estudiante in enumerate(archivos_estudiantes, 1):
        nombre = os.path.basename(ruta_estudiante)
        logger.info(f"\n[{idx}/{total_archivos}] Procesando: {nombre}")

        # Agregar sufijo _REV al nombre del archivo revisado
        nombre_sin_ext, extension = os.path.splitext(nombre)
        nombre_revisado = f"{nombre_sin_ext}_REV{extension}"
        ruta_salida = os.path.join(args.salida, nombre_revisado)

        try:
            resultado = auditar_libro(args.plantilla, ruta_estudiante, ruta_salida)
            resultados.append(resultado)
        except Exception as e:
            logger.error(f"❌ Error inesperado procesando '{nombre}': {e}")
            resultados.append({
                "archivo": nombre,
                "total_aciertos": 0,
                "total_errores": -1,
                "detalle_hojas": [],
                "detalles_com": [f"Error fatal: {e}"]
            })

    # ----------------------------------------------------------------
    # GENERAR LOG XLSX
    # ----------------------------------------------------------------
    generar_log_xlsx(resultados, args.log, args.plantilla, args.trabajos)

    # ----------------------------------------------------------------
    # RESUMEN FINAL
    # ----------------------------------------------------------------
    logger.info("\n" + "=" * 70)
    logger.info("  📊 RESUMEN FINAL DE AUDITORÍA")
    logger.info("=" * 70)
    logger.info(f"  Archivos procesados: {total_archivos}")

    total_aciertos_global = sum(r["total_aciertos"] for r in resultados)
    total_errores_global = sum(
        r["total_errores"] for r in resultados if r["total_errores"] >= 0
    )

    logger.info(f"  Total aciertos (global):  {total_aciertos_global}")
    logger.info(f"  Total errores (global):   {total_errores_global}")
    logger.info(f"\n  📁 Archivos revisados en: {args.salida}")
    logger.info(f"  📄 Log de notas en:       {args.log}")
    logger.info("=" * 70)

    # Tabla resumen por estudiante
    logger.info(f"\n{'Estudiante':<40} {'Aciertos':>10} {'Errores':>10} {'%':>8}")
    logger.info("-" * 70)
    for res in resultados:
        nombre = os.path.splitext(res["archivo"])[0]
        ac = res["total_aciertos"]
        er = res["total_errores"]
        total = ac + er if er >= 0 else 0
        pct = (ac / total * 100) if total > 0 else 0
        estado = "✅" if er == 0 else "⚠️" if er < 5 else "❌"
        logger.info(f"  {estado} {nombre:<37} {ac:>10} {er:>10} {pct:>7.1f}%")

    logger.info("\n✅ Auditoría completada.")


if __name__ == "__main__":
    main()
