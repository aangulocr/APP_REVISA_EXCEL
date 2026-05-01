# ============================================================================
# auditor.py — Motor principal de auditoría
# ============================================================================
"""
Orquesta la comparación completa entre un libro plantilla y un libro
de estudiante. Coordina los módulos de comparación de celdas, objetos
y COM para producir un resultado consolidado.
"""

import os
import logging
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

from comparador_celdas import comparar_celda, marcar_celda_con_error
from comparador_objetos import (
    comparar_tablas,
    comparar_celdas_combinadas,
    comparar_validaciones,
    comparar_formato_condicional,
    comparar_dimensiones,
    comparar_graficos_openpyxl,
)
from comparador_com import inspeccion_profunda_com
from config import MENSAJES

logger = logging.getLogger(__name__)


def _obtener_rango_usado(ws):
    """
    Determina el rango de celdas usado en una hoja.
    Retorna (max_row, max_col).
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    return max_row, max_col


def auditar_hoja(ws_plantilla, ws_estudiante, nombre_hoja):
    """
    Realiza la auditoría celda a celda y de objetos complejos
    para una hoja específica.

    Parámetros:
        ws_plantilla: Worksheet de la plantilla (data_only=False).
        ws_estudiante: Worksheet del estudiante.
        nombre_hoja: Nombre de la hoja (para el log).

    Retorna:
        dict: {
            "hoja": str,
            "aciertos": int,
            "errores": int,
            "detalles_objetos": list[str]
        }
    """
    aciertos = 0
    errores = 0
    detalles_objetos = []

    # ----------------------------------------------------------------
    # 1. COMPARACIÓN CELDA A CELDA
    # ----------------------------------------------------------------
    max_row_p, max_col_p = _obtener_rango_usado(ws_plantilla)
    max_row_e, max_col_e = _obtener_rango_usado(ws_estudiante)

    # Usar el rango más grande para no dejar celdas sin revisar
    max_row = max(max_row_p, max_row_e)
    max_col = max(max_col_p, max_col_e)

    logger.info(
        f"  📋 Hoja '{nombre_hoja}': Analizando {max_row} filas × {max_col} columnas..."
    )

    for fila in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            coord = f"{col_letter}{fila}"

            celda_p = ws_plantilla[coord]
            celda_e = ws_estudiante[coord]

            # Saltar celdas vacías en ambos archivos
            if celda_p.value is None and celda_e.value is None:
                # Aún así comparar formato si la plantilla tiene formato especial
                if (celda_p.fill.fill_type is not None or
                    celda_p.border.left.style is not None or
                    celda_p.border.right.style is not None or
                    celda_p.border.top.style is not None or
                    celda_p.border.bottom.style is not None):
                    es_correcto, lista_errores = comparar_celda(celda_p, celda_e)
                    if es_correcto:
                        aciertos += 1
                    else:
                        errores += len(lista_errores)
                        marcar_celda_con_error(celda_e, lista_errores)
                continue

            es_correcto, lista_errores = comparar_celda(celda_p, celda_e)
            if es_correcto:
                aciertos += 1
            else:
                errores += len(lista_errores)
                marcar_celda_con_error(celda_e, lista_errores)

    # ----------------------------------------------------------------
    # 2. COMPARACIÓN DE TABLAS DE EXCEL
    # ----------------------------------------------------------------
    try:
        err_tablas, det_tablas = comparar_tablas(ws_plantilla, ws_estudiante)
        errores += err_tablas
        detalles_objetos.extend(det_tablas)
    except Exception as e:
        detalles_objetos.append(f"⚠️ Error comparando tablas: {e}")

    # ----------------------------------------------------------------
    # 3. COMPARACIÓN DE CELDAS COMBINADAS
    # ----------------------------------------------------------------
    try:
        err_merge, det_merge = comparar_celdas_combinadas(ws_plantilla, ws_estudiante)
        errores += err_merge
        detalles_objetos.extend(det_merge)
    except Exception as e:
        detalles_objetos.append(f"⚠️ Error comparando celdas combinadas: {e}")

    # ----------------------------------------------------------------
    # 4. COMPARACIÓN DE VALIDACIONES DE DATOS
    # ----------------------------------------------------------------
    try:
        err_val, det_val = comparar_validaciones(ws_plantilla, ws_estudiante)
        errores += err_val
        detalles_objetos.extend(det_val)
    except Exception as e:
        detalles_objetos.append(f"⚠️ Error comparando validaciones: {e}")

    # ----------------------------------------------------------------
    # 5. COMPARACIÓN DE FORMATO CONDICIONAL
    # ----------------------------------------------------------------
    try:
        err_fc, det_fc = comparar_formato_condicional(ws_plantilla, ws_estudiante)
        errores += err_fc
        detalles_objetos.extend(det_fc)
    except Exception as e:
        detalles_objetos.append(f"⚠️ Error comparando formato condicional: {e}")

    # ----------------------------------------------------------------
    # 6. COMPARACIÓN DE DIMENSIONES
    # ----------------------------------------------------------------
    try:
        err_dim, det_dim = comparar_dimensiones(ws_plantilla, ws_estudiante)
        errores += err_dim
        detalles_objetos.extend(det_dim)
    except Exception as e:
        detalles_objetos.append(f"⚠️ Error comparando dimensiones: {e}")

    # ----------------------------------------------------------------
    # 7. COMPARACIÓN DE GRÁFICOS (openpyxl)
    # ----------------------------------------------------------------
    try:
        err_chart, det_chart = comparar_graficos_openpyxl(ws_plantilla, ws_estudiante)
        errores += err_chart
        detalles_objetos.extend(det_chart)
    except Exception as e:
        detalles_objetos.append(f"⚠️ Error comparando gráficos: {e}")

    return {
        "hoja": nombre_hoja,
        "aciertos": aciertos,
        "errores": errores,
        "detalles_objetos": detalles_objetos
    }


def auditar_libro(ruta_plantilla, ruta_estudiante, ruta_salida):
    """
    Audita un libro de estudiante contra la plantilla.

    Parámetros:
        ruta_plantilla: Ruta al archivo PLANTILLA.xlsx.
        ruta_estudiante: Ruta al archivo del estudiante .xlsx.
        ruta_salida: Ruta donde guardar el archivo revisado.

    Retorna:
        dict: {
            "archivo": str,
            "total_aciertos": int,
            "total_errores": int,
            "detalle_hojas": list[dict],
            "detalles_com": list[str]
        }
    """
    nombre_archivo = os.path.basename(ruta_estudiante)
    logger.info(f"\n{'='*70}")
    logger.info(f"📖 Auditando: {nombre_archivo}")
    logger.info(f"{'='*70}")

    resultado = {
        "archivo": nombre_archivo,
        "total_aciertos": 0,
        "total_errores": 0,
        "detalle_hojas": [],
        "detalles_com": []
    }

    try:
        # Abrir la plantilla (con fórmulas)
        wb_plantilla = openpyxl.load_workbook(ruta_plantilla, data_only=False)

        # Abrir el libro del estudiante (con fórmulas, para editar y guardar)
        wb_estudiante = openpyxl.load_workbook(ruta_estudiante, data_only=False)

    except Exception as e:
        logger.error(f"❌ No se pudo abrir el archivo '{nombre_archivo}': {e}")
        resultado["total_errores"] = -1  # Indicador de fallo de apertura
        resultado["detalles_com"].append(f"Error de apertura: {e}")
        return resultado

    # ----------------------------------------------------------------
    # COMPARACIÓN POR HOJAS
    # ----------------------------------------------------------------
    hojas_plantilla = wb_plantilla.sheetnames

    for nombre_hoja in hojas_plantilla:
        if nombre_hoja in wb_estudiante.sheetnames:
            ws_p = wb_plantilla[nombre_hoja]
            ws_e = wb_estudiante[nombre_hoja]

            res_hoja = auditar_hoja(ws_p, ws_e, nombre_hoja)

            resultado["total_aciertos"] += res_hoja["aciertos"]
            resultado["total_errores"] += res_hoja["errores"]
            resultado["detalle_hojas"].append(res_hoja)

            logger.info(
                f"  ✅ Hoja '{nombre_hoja}': "
                f"{res_hoja['aciertos']} aciertos | "
                f"{res_hoja['errores']} errores"
            )
            if res_hoja["detalles_objetos"]:
                for detalle in res_hoja["detalles_objetos"]:
                    logger.info(f"     {detalle}")
        else:
            # La hoja no existe en el libro del estudiante
            msg = MENSAJES["hoja_faltante"].format(nombre=nombre_hoja)
            logger.warning(f"  ⚠️ {msg}")
            resultado["total_errores"] += 1
            resultado["detalle_hojas"].append({
                "hoja": nombre_hoja,
                "aciertos": 0,
                "errores": 1,
                "detalles_objetos": [msg]
            })

    # ----------------------------------------------------------------
    # INSPECCIÓN PROFUNDA CON COM (Tablas Dinámicas + Gráficos)
    # ----------------------------------------------------------------
    try:
        err_com, det_com = inspeccion_profunda_com(ruta_plantilla, ruta_estudiante)
        resultado["total_errores"] += err_com
        resultado["detalles_com"] = det_com

        if det_com:
            logger.info(f"\n  🔍 Inspección COM (Tablas Dinámicas / Gráficos):")
            for detalle in det_com:
                logger.info(f"     {detalle}")
    except Exception as e:
        logger.warning(f"  ⚠️ Error en inspección COM: {e}")
        resultado["detalles_com"].append(f"Error COM: {e}")

    # ----------------------------------------------------------------
    # GUARDAR ARCHIVO REVISADO
    # ----------------------------------------------------------------
    try:
        wb_estudiante.save(ruta_salida)
        logger.info(f"  💾 Guardado en: {ruta_salida}")
    except Exception as e:
        logger.error(f"  ❌ No se pudo guardar '{ruta_salida}': {e}")

    wb_plantilla.close()
    wb_estudiante.close()

    return resultado
