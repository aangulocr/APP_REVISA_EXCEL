# ============================================================================
# crear_datos_prueba.py — Genera archivos de prueba para validar el auditor
# ============================================================================
"""
Crea una PLANTILLA.xlsx y varios archivos de estudiantes con diferencias
intencionales para probar todas las capacidades del auditor.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANTILLA_PATH = os.path.join(BASE_DIR, "PLANTILLA.xlsx")
TRABAJOS_DIR = os.path.join(BASE_DIR, "TRABAJOS_ESTUDIANTES")


def crear_plantilla():
    """Crea la PLANTILLA.xlsx con formatos, fórmulas, validaciones, etc."""
    wb = Workbook()

    # ============================
    # HOJA 1: Datos y Fórmulas
    # ============================
    ws1 = wb.active
    ws1.title = "Ventas"

    # Encabezados con formato
    encabezados = ["Producto", "Cantidad", "Precio Unitario", "Total"]
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, texto in enumerate(encabezados, 1):
        celda = ws1.cell(row=1, column=col, value=texto)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_align
        celda.border = thin_border

    # Datos
    datos = [
        ("Laptop", 10, 899.99),
        ("Mouse", 50, 25.50),
        ("Teclado", 30, 45.00),
        ("Monitor", 15, 350.00),
        ("Audífonos", 40, 75.00),
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")

    for row_idx, (producto, cantidad, precio) in enumerate(datos, 2):
        ws1.cell(row=row_idx, column=1, value=producto).font = data_font
        ws1.cell(row=row_idx, column=1).alignment = data_align
        ws1.cell(row=row_idx, column=1).border = thin_border

        ws1.cell(row=row_idx, column=2, value=cantidad).font = data_font
        ws1.cell(row=row_idx, column=2).alignment = num_align
        ws1.cell(row=row_idx, column=2).border = thin_border
        ws1.cell(row=row_idx, column=2).number_format = '#,##0'

        ws1.cell(row=row_idx, column=3, value=precio).font = data_font
        ws1.cell(row=row_idx, column=3).alignment = num_align
        ws1.cell(row=row_idx, column=3).border = thin_border
        ws1.cell(row=row_idx, column=3).number_format = '$#,##0.00'

        # Fórmula: Total = Cantidad * Precio
        ws1.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}").font = data_font
        ws1.cell(row=row_idx, column=4).alignment = num_align
        ws1.cell(row=row_idx, column=4).border = thin_border
        ws1.cell(row=row_idx, column=4).number_format = '$#,##0.00'

    # Fila de totales
    row_total = len(datos) + 2
    ws1.cell(row=row_total, column=1, value="TOTAL").font = Font(name="Calibri", size=12, bold=True)
    ws1.cell(row=row_total, column=1).border = thin_border
    ws1.cell(row=row_total, column=2, value=f"=SUM(B2:B{row_total-1})").font = Font(name="Calibri", size=11, bold=True)
    ws1.cell(row=row_total, column=2).border = thin_border
    ws1.cell(row=row_total, column=2).number_format = '#,##0'
    ws1.cell(row=row_total, column=4, value=f"=SUM(D2:D{row_total-1})").font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    ws1.cell(row=row_total, column=4).border = thin_border
    ws1.cell(row=row_total, column=4).number_format = '$#,##0.00'

    # Anchos de columna
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18

    # Validación de datos: Cantidad debe ser entre 1 y 1000
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=1000)
    dv.error = "La cantidad debe estar entre 1 y 1000"
    dv.errorTitle = "Valor inválido"
    ws1.add_data_validation(dv)
    dv.add(f"B2:B{row_total-1}")

    # Formato condicional: resaltar totales > 5000
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE")
    red_font = Font(color="9C0006")
    ws1.conditional_formatting.add(
        f"D2:D{row_total-1}",
        CellIsRule(operator="greaterThan", formula=["5000"], fill=red_fill, font=red_font)
    )

    # Celdas combinadas
    ws1.merge_cells("A9:D9")
    ws1.cell(row=9, column=1, value="Reporte de Ventas - 2026").font = Font(name="Calibri", size=14, bold=True, italic=True)
    ws1.cell(row=9, column=1).alignment = Alignment(horizontal="center")

    # Gráfico
    chart = BarChart()
    chart.title = "Ventas por Producto"
    chart.x_axis.title = "Producto"
    chart.y_axis.title = "Total ($)"
    chart.style = 10

    data_ref = Reference(ws1, min_col=4, min_row=1, max_row=row_total-1)
    cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=row_total-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws1.add_chart(chart, "F2")

    # ============================
    # HOJA 2: Resumen
    # ============================
    ws2 = wb.create_sheet("Resumen")
    ws2.cell(row=1, column=1, value="Estadísticas").font = Font(name="Calibri", size=14, bold=True, color="4472C4")
    ws2.cell(row=2, column=1, value="Promedio de ventas:")
    ws2.cell(row=2, column=2, value="=AVERAGE(Ventas!D2:D6)").number_format = '$#,##0.00'
    ws2.cell(row=3, column=1, value="Venta máxima:")
    ws2.cell(row=3, column=2, value="=MAX(Ventas!D2:D6)").number_format = '$#,##0.00'
    ws2.cell(row=4, column=1, value="Venta mínima:")
    ws2.cell(row=4, column=2, value="=MIN(Ventas!D2:D6)").number_format = '$#,##0.00'
    ws2.cell(row=5, column=1, value="Total productos:")
    ws2.cell(row=5, column=2, value="=COUNTA(Ventas!A2:A6)")

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18

    wb.save(PLANTILLA_PATH)
    print(f"PLANTILLA creada: {PLANTILLA_PATH}")


def crear_estudiante_perfecto():
    """Crea un archivo de estudiante PERFECTO (sin errores)."""
    # Simplemente copiar la plantilla
    import shutil
    os.makedirs(TRABAJOS_DIR, exist_ok=True)
    destino = os.path.join(TRABAJOS_DIR, "Estudiante_Perfecto.xlsx")
    shutil.copy2(PLANTILLA_PATH, destino)
    print(f"Estudiante perfecto creado: {destino}")


def crear_estudiante_con_errores():
    """Crea un archivo con ERRORES INTENCIONALES en varias categorías."""
    wb = Workbook()

    # ============================
    # HOJA 1: Ventas (con errores)
    # ============================
    ws1 = wb.active
    ws1.title = "Ventas"

    # Encabezados — ERROR: fuente diferente y sin relleno
    encabezados = ["Producto", "Cantidad", "Precio Unitario", "Total"]
    wrong_font = Font(name="Arial", size=10, bold=False, color="000000")  # Fuente incorrecta
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, texto in enumerate(encabezados, 1):
        celda = ws1.cell(row=1, column=col, value=texto)
        celda.font = wrong_font  # ERROR: fuente incorrecta
        # ERROR: sin relleno (falta fill azul)
        celda.border = thin_border

    # Datos — algunos valores incorrectos
    datos = [
        ("Laptop", 10, 899.99),
        ("Mouse", 45, 25.50),        # ERROR: Cantidad cambiada (50 -> 45)
        ("Teclado", 30, 45.00),
        ("Pantalla", 15, 350.00),     # ERROR: Nombre cambiado (Monitor -> Pantalla)
        ("Audífonos", 40, 75.00),
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")

    for row_idx, (producto, cantidad, precio) in enumerate(datos, 2):
        ws1.cell(row=row_idx, column=1, value=producto).font = data_font
        ws1.cell(row=row_idx, column=1).alignment = data_align
        ws1.cell(row=row_idx, column=1).border = thin_border

        ws1.cell(row=row_idx, column=2, value=cantidad).font = data_font
        ws1.cell(row=row_idx, column=2).alignment = num_align
        ws1.cell(row=row_idx, column=2).border = thin_border
        ws1.cell(row=row_idx, column=2).number_format = '#,##0'

        ws1.cell(row=row_idx, column=3, value=precio).font = data_font
        ws1.cell(row=row_idx, column=3).alignment = num_align
        ws1.cell(row=row_idx, column=3).border = thin_border
        ws1.cell(row=row_idx, column=3).number_format = '$#,##0.00'

        # ERROR en fila 4: fórmula diferente (usa + en vez de *)
        if row_idx == 4:
            ws1.cell(row=row_idx, column=4, value=f"=B{row_idx}+C{row_idx}").font = data_font
        else:
            ws1.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}").font = data_font

        ws1.cell(row=row_idx, column=4).alignment = num_align
        ws1.cell(row=row_idx, column=4).border = thin_border
        ws1.cell(row=row_idx, column=4).number_format = '$#,##0.00'

    # Fila totales — ERROR: usa AVERAGE en vez de SUMA
    row_total = 7
    ws1.cell(row=row_total, column=1, value="TOTAL").font = Font(name="Calibri", size=12, bold=True)
    ws1.cell(row=row_total, column=1).border = thin_border
    ws1.cell(row=row_total, column=2, value=f"=SUM(B2:B{row_total-1})").font = Font(name="Calibri", size=11, bold=True)
    ws1.cell(row=row_total, column=2).border = thin_border
    ws1.cell(row=row_total, column=2).number_format = '#,##0'
    # ERROR: usa AVERAGE en vez de SUM
    ws1.cell(row=row_total, column=4, value=f"=AVERAGE(D2:D{row_total-1})").font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    ws1.cell(row=row_total, column=4).border = thin_border
    ws1.cell(row=row_total, column=4).number_format = '$#,##0.00'

    # Anchos de columna — ERROR: diferentes
    ws1.column_dimensions['A'].width = 15  # ERROR: 15 vs 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18

    # ERROR: sin validación de datos

    # ERROR: sin formato condicional

    # ERROR: sin celdas combinadas en A9:D9

    # ERROR: sin gráfico

    # ============================
    # HOJA 2: Resumen (con errores)
    # ============================
    ws2 = wb.create_sheet("Resumen")
    ws2.cell(row=1, column=1, value="Estadísticas").font = Font(name="Arial", size=14, bold=True, color="000000")  # ERROR: fuente y color
    ws2.cell(row=2, column=1, value="Promedio de ventas:")
    ws2.cell(row=2, column=2, value="=AVERAGE(Ventas!D2:D6)").number_format = '$#,##0.00'
    ws2.cell(row=3, column=1, value="Venta máxima:")
    ws2.cell(row=3, column=2, value="=MAX(Ventas!D2:D6)").number_format = '$#,##0.00'
    # ERROR: falta la fila 4 (Venta mínima)
    ws2.cell(row=5, column=1, value="Total productos:")
    ws2.cell(row=5, column=2, value="=COUNTA(Ventas!A2:A6)")

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18

    os.makedirs(TRABAJOS_DIR, exist_ok=True)
    ruta = os.path.join(TRABAJOS_DIR, "Estudiante_Con_Errores.xlsx")
    wb.save(ruta)
    print(f"Estudiante con errores creado: {ruta}")


def crear_estudiante_hoja_faltante():
    """Crea un archivo al que le FALTA una hoja completa."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Solo la hoja Ventas, falta "Resumen"
    ws.cell(row=1, column=1, value="Producto").font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=2, column=1, value="Laptop")

    os.makedirs(TRABAJOS_DIR, exist_ok=True)
    ruta = os.path.join(TRABAJOS_DIR, "Estudiante_Hoja_Faltante.xlsx")
    wb.save(ruta)
    print(f"Estudiante con hoja faltante creado: {ruta}")


if __name__ == "__main__":
    print("Generando datos de prueba...")
    crear_plantilla()
    crear_estudiante_perfecto()
    crear_estudiante_con_errores()
    crear_estudiante_hoja_faltante()
    print("\nDatos de prueba generados exitosamente.")
    print(f"  Plantilla:  {PLANTILLA_PATH}")
    print(f"  Trabajos:   {TRABAJOS_DIR}")
